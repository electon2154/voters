from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q, Count
from .models import CustomUser, Entity, Candidate, Pillar, Voter, AppearanceSettings
from .forms import LoginForm, ExcelUploadForm, VoterForm, PillarForm, CandidateForm, VoterCandidateForm, EntityForm
import json
import openpyxl

# صفحة تسجيل الدخول
def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                # توجيه المستخدم حسب نوعه
                if user.user_type == 'entity':
                    return redirect('elections:entity_dashboard')
                elif user.user_type == 'candidate':
                    return redirect('elections:candidate_dashboard')
                elif user.user_type == 'pillar':
                    return redirect('elections:pillar_dashboard')
                elif user.user_type == 'admin':
                    return redirect('elections:admin_dashboard')
            else:
                messages.error(request, 'اسم المستخدم أو كلمة المرور غير صحيحة')
    else:
        form = LoginForm()
    return render(request, 'elections/login.html', {'form': form})

# تسجيل الخروج
def logout_view(request):
    logout(request)
    return redirect('elections:login')

# لوحة تحكم الكيان
@login_required
def entity_dashboard(request):
    if request.user.user_type != 'entity':
        return redirect('elections:login')
    
    entity = get_object_or_404(Entity, user=request.user)
    candidates = entity.candidates.all()
    
    # إحصائيات عامة
    total_candidates = candidates.count()
    total_pillars = Pillar.objects.filter(candidate__entity=entity).count()
    total_voters = Voter.objects.filter(candidate__entity=entity).count()
    total_centers = Voter.objects.filter(candidate__entity=entity).values('center_number').distinct().count()
    total_stations = Voter.objects.filter(candidate__entity=entity).values('station').distinct().count()
    updated_voters = Voter.objects.filter(candidate__entity=entity, card_status='updated').count()
    not_updated_voters = Voter.objects.filter(candidate__entity=entity, card_status='not_updated').count()
    voted_voters = Voter.objects.filter(candidate__entity=entity, voting_status='voted').count()
    not_voted_voters = Voter.objects.filter(candidate__entity=entity, voting_status='not_voted').count()
    
    # إحصائيات لكل مرشح
    candidate_stats = []
    for candidate in candidates:
        candidate_voters = candidate.voters.count()
        candidate_voted = candidate.voters.filter(voting_status='voted').count()
        candidate_not_voted = candidate.voters.filter(voting_status='not_voted').count()
        candidate_pillars = candidate.pillars.count()
        voting_percentage = (candidate_voted / candidate_voters * 100) if candidate_voters > 0 else 0
        candidate_stats.append({
            'candidate': candidate,
            'voters_count': candidate_voters,
            'voted_count': candidate_voted,
            'not_voted_count': candidate_not_voted,
            'pillars_count': candidate_pillars,
            'voting_percentage': round(voting_percentage, 1)
        })
    
    # تجميع الإحصائيات في متغير stats
    stats = {
        'total_candidates': total_candidates,
        'total_pillars': total_pillars,
        'total_voters': total_voters,
        'total_centers': total_centers,
        'total_stations': total_stations,
        'updated_cards': updated_voters,
        'not_updated_cards': not_updated_voters,
        'voted_voters': voted_voters,
        'not_voted_voters': not_voted_voters,
    }
    
    context = {
        'entity': entity,
        'stats': stats,
        'candidate_stats': candidate_stats,
    }
    return render(request, 'elections/entity_dashboard.html', context)

# لوحة تحكم المرشح
@login_required
def candidate_dashboard(request):
    if request.user.user_type != 'candidate':
        return redirect('elections:login')
    
    candidate = get_object_or_404(Candidate, user=request.user)
    pillars = candidate.pillars.all()
    
    # إحصائيات
    total_pillars = pillars.count()
    total_voters = candidate.voters.count()
    total_centers = candidate.voters.values('center_number').distinct().count()
    total_stations = candidate.voters.values('station').distinct().count()
    updated_voters = candidate.voters.filter(card_status='updated').count()
    not_updated_voters = candidate.voters.filter(card_status='not_updated').count()
    voted_voters = candidate.voters.filter(voting_status='voted').count()
    not_voted_voters = candidate.voters.filter(voting_status='not_voted').count()
    
    context = {
        'candidate': candidate,
        'pillars': pillars,
        'stats': {
            'total_pillars': total_pillars,
            'total_voters': total_voters,
            'total_centers': total_centers,
            'total_stations': total_stations,
            'updated_cards': updated_voters,
            'not_updated_cards': not_updated_voters,
            'voted': voted_voters,
            'not_voted': not_voted_voters,
        }
    }
    return render(request, 'elections/candidate_dashboard.html', context)

# لوحة تحكم الركيزة
@login_required
def pillar_dashboard(request):
    if request.user.user_type != 'pillar':
        return redirect('elections:login')
    
    pillar = get_object_or_404(Pillar, user=request.user)
    voters = pillar.voters.all()
    
    # البحث والفلترة
    search_query = request.GET.get('search', '')
    card_status_filter = request.GET.get('card_status', '')
    voting_status_filter = request.GET.get('voting_status', '')
    district_filter = request.GET.get('district', '')
    sub_district_filter = request.GET.get('sub_district', '')
    
    if search_query:
        voters = voters.filter(
            Q(name__icontains=search_query) |
            Q(voter_number__icontains=search_query) |
            Q(phone_number__icontains=search_query)
        )
    
    if card_status_filter:
        voters = voters.filter(card_status=card_status_filter)
    
    if voting_status_filter:
        voters = voters.filter(voting_status=voting_status_filter)
    
    if district_filter:
        voters = voters.filter(district=district_filter)
    
    if sub_district_filter:
        voters = voters.filter(sub_district=sub_district_filter)
    
    # التصفح
    paginator = Paginator(voters, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # إحصائيات
    total_voters = pillar.voters.count()
    total_centers = pillar.voters.values('center_number').distinct().count()
    total_stations = pillar.voters.values('station').distinct().count()
    updated_voters = pillar.voters.filter(card_status='updated').count()
    not_updated_voters = pillar.voters.filter(card_status='not_updated').count()
    voted_voters = pillar.voters.filter(voting_status='voted').count()
    not_voted_voters = pillar.voters.filter(voting_status='not_voted').count()
    
    # حساب النسب المئوية
    voting_percentage = (voted_voters / total_voters * 100) if total_voters > 0 else 0
    update_percentage = (updated_voters / total_voters * 100) if total_voters > 0 else 0
    
    stats = {
        'total_voters': total_voters,
        'total_centers': total_centers,
        'total_stations': total_stations,
        'updated_cards': updated_voters,
        'not_updated_cards': not_updated_voters,
        'voted': voted_voters,
        'not_voted': not_voted_voters,
        'voting_percentage': voting_percentage,
        'update_percentage': update_percentage,
    }
    
    # الحصول على قوائم المناطق والنواحي المتاحة
    governorates = pillar.voters.values_list('governorate', flat=True).distinct().order_by('governorate')
    districts = pillar.voters.values_list('district', flat=True).distinct().order_by('district')
    sub_districts = pillar.voters.values_list('sub_district', flat=True).distinct().order_by('sub_district')
    
    context = {
        'pillar': pillar,
        'voters': page_obj,
        'stats': stats,
        'search_query': search_query,
        'card_status_filter': card_status_filter,
        'voting_status_filter': voting_status_filter,
        'district_filter': district_filter,
        'sub_district_filter': sub_district_filter,
        'governorates': governorates,
        'districts': districts,
        'sub_districts': sub_districts,
    }
    return render(request, 'elections/pillar_dashboard.html', context)

# تحديث حالة التصويت
@login_required
def update_voter_status(request, voter_id):
    if request.user.user_type != 'pillar':
        return JsonResponse({'success': False, 'error': 'غير مصرح'})
    
    pillar = get_object_or_404(Pillar, user=request.user)
    voter = get_object_or_404(Voter, id=voter_id, pillar=pillar)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        new_status = data.get('status')
        
        # تحويل القيم العربية إلى القيم المخزنة في قاعدة البيانات
        status_mapping = {
            'صوت': 'voted',
            'لم يصوت': 'not_voted'
        }
        
        if new_status in status_mapping:
            voter.voting_status = status_mapping[new_status]
            voter.save()
            return JsonResponse({'success': True})
    
    return JsonResponse({'success': False})

# إضافة مرشح جديد (للكيان)
@login_required
def add_candidate(request):
    if request.user.user_type != 'entity':
        return redirect('elections:login')
    
    entity = get_object_or_404(Entity, user=request.user)
    
    if request.method == 'POST':
        form = CandidateForm(request.POST, request.FILES)
        if form.is_valid():
            candidate = form.save(entity)
            messages.success(request, 'تم إضافة المرشح بنجاح')
            return redirect('elections:entity_dashboard')
    else:
        form = CandidateForm()
    
    return render(request, 'elections/add_candidate.html', {'form': form})

# إضافة ركيزة جديدة (للمرشح)
@login_required
def add_pillar(request):
    if request.user.user_type != 'candidate':
        return redirect('login')
    
    candidate = get_object_or_404(Candidate, user=request.user)
    
    if request.method == 'POST':
        form = PillarForm(request.POST)
        if form.is_valid():
            pillar = form.save(candidate)
            messages.success(request, 'تم إضافة الركيزة بنجاح')
            return redirect('elections:candidate_dashboard')
    else:
        form = PillarForm()
    
    return render(request, 'elections/add_pillar.html', {'form': form})

# إضافة ناخب جديد (للركيزة)
@login_required
def add_voter(request):
    if request.user.user_type != 'pillar':
        return redirect('elections:login')
    
    pillar = get_object_or_404(Pillar, user=request.user)
    
    if request.method == 'POST':
        form = VoterForm(request.POST)
        if form.is_valid():
            voter = form.save(commit=False)
            voter.pillar = pillar
            voter.candidate = pillar.candidate  # ضبط المرشح تلقائياً
            voter.voting_status = 'not_voted'  # ضبط حالة التصويت تلقائياً
            voter.save()
            messages.success(request, 'تم إضافة الناخب بنجاح')
            return redirect('elections:pillar_dashboard')
    else:
        form = VoterForm()
    
    return render(request, 'elections/add_voter.html', {'form': form})

# إضافة ناخبين للكيان
@login_required
def add_voters_to_entity(request):
    if request.user.user_type != 'entity':
        return redirect('elections:login')
    
    entity = get_object_or_404(Entity, user=request.user)
    candidates = entity.candidates.all()
    
    if request.method == 'POST':
        # الحصول على البيانات من النموذج
        candidate_id = request.POST.get('candidate')
        pillar_id = request.POST.get('pillar')
        voter_number = request.POST.get('voter_number')
        name = request.POST.get('name')
        governorate = request.POST.get('governorate')
        district = request.POST.get('district')
        sub_district = request.POST.get('sub_district')
        card_status = request.POST.get('card_status')
        center_name = request.POST.get('center_name')
        center_number = request.POST.get('center_number')
        station = request.POST.get('station')
        phone_number = request.POST.get('phone_number')
        
        # التحقق من صحة البيانات
        if candidate_id and pillar_id and voter_number and name:
            try:
                candidate = get_object_or_404(Candidate, id=candidate_id, entity=entity)
                pillar = get_object_or_404(Pillar, id=pillar_id, candidate=candidate)
                
                # إنشاء الناخب الجديد
                voter = Voter.objects.create(
                    voter_number=voter_number,
                    name=name,
                    governorate=governorate,
                    district=district,
                    sub_district=sub_district or '',
                    card_status=card_status,
                    center_name=center_name,
                    center_number=center_number,
                    station=station,
                    phone_number=phone_number,
                    pillar=pillar,
                    candidate=candidate,
                    voting_status='not_voted'
                )
                
                messages.success(request, 'تم إضافة الناخب بنجاح')
                return redirect('elections:entity_dashboard')
                
            except Exception as e:
                messages.error(request, f'حدث خطأ أثناء إضافة الناخب: {str(e)}')
        else:
            messages.error(request, 'يرجى ملء جميع الحقول المطلوبة')
    
    context = {
        'candidates': candidates
    }
    return render(request, 'elections/add_voters_entity.html', context)

# إضافة ناخبين للمرشح
@login_required
def add_voters_to_candidate(request):
    if request.user.user_type != 'candidate':
        return redirect('elections:login')
    
    candidate = get_object_or_404(Candidate, user=request.user)
    
    if request.method == 'POST':
        form = VoterCandidateForm(request.POST, candidate=candidate)
        if form.is_valid():
            voter = form.save(commit=False)
            voter.candidate = candidate
            voter.voting_status = 'not_voted'  # ضبط حالة التصويت تلقائياً
            
            # تحديد الركيزة بناءً على الاختيار
            pillar = form.cleaned_data['pillar']
            voter.pillar = pillar
            
            voter.save()
            messages.success(request, 'تم إضافة الناخب بنجاح')
            return redirect('elections:candidate_dashboard')
    else:
        form = VoterCandidateForm(candidate=candidate)
    
    return render(request, 'elections/add_voters_candidate.html', {'form': form})

# رفع ملف Excel للكيان
@login_required
def upload_excel_entity(request):
    if request.user.user_type != 'entity':
        return redirect('elections:login')
    
    if request.method == 'POST':
        try:
            if 'excel_file' not in request.FILES:
                return JsonResponse({'success': False, 'error': 'لم يتم اختيار ملف'})
            
            excel_file = request.FILES['excel_file']
            
            # التحقق من نوع الملف
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({'success': False, 'error': 'يجب أن يكون الملف من نوع Excel'})
            
            # قراءة ملف Excel
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            # معالجة البيانات وإضافتها للقاعدة
            created_count = 0
            errors = []
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # إنشاء مستخدم جديد للكيان
                    user = CustomUser.objects.create_user(
                        username=row[0] if row[0] else f'entity_{row_num}',
                        email=row[1] if row[1] else '',
                        first_name=row[2] if row[2] else '',
                        last_name=row[3] if row[3] else '',
                        password='defaultpassword123',
                        user_type='entity'
                    )
                    
                    # إنشاء كيان جديد
                    entity = Entity.objects.create(
                        user=user,
                        phone=row[4] if row[4] else '',
                        address=row[5] if row[5] else ''
                    )
                    
                    created_count += 1
                except Exception as e:
                    errors.append(f'الصف {row_num}: {str(e)}')
            
            return JsonResponse({
                'success': True, 
                'message': f'تم إنشاء {created_count} كيان بنجاح',
                'errors': errors
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'خطأ في معالجة الملف: {str(e)}'})
    
    return render(request, 'elections/upload_excel.html', {'type': 'entity', 'title': 'رفع ملف الكيانات'})

# رفع ملف Excel للمرشح
@login_required
def upload_excel_candidate(request):
    if request.user.user_type != 'candidate':
        return redirect('elections:login')
    
    candidate = get_object_or_404(Candidate, user=request.user)
    
    if request.method == 'POST':
        try:
            if 'excel_file' not in request.FILES:
                return JsonResponse({'success': False, 'error': 'لم يتم اختيار ملف'})
            
            excel_file = request.FILES['excel_file']
            
            # التحقق من نوع الملف
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                return JsonResponse({'success': False, 'error': 'يجب أن يكون الملف من نوع Excel'})
            
            # قراءة ملف Excel
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            
            # معالجة البيانات وإضافتها للقاعدة
            created_count = 0
            errors = []
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # إنشاء ناخب جديد
                    voter = Voter.objects.create(
                        name=row[0] if row[0] else '',
                        voter_number=row[1] if row[1] else '',
                        phone_number=row[2] if row[2] else '',
                        address=row[3] if row[3] else '',
                        center_number=row[4] if row[4] else '',
                        station=row[5] if row[5] else '',
                        card_status=row[6] if row[6] else 'not_updated',
                        voting_status=row[7] if row[7] else 'not_voted',
                        notes=row[8] if row[8] else '',
                        candidate=candidate
                    )
                    
                    created_count += 1
                except Exception as e:
                    errors.append(f'الصف {row_num}: {str(e)}')
            
            return JsonResponse({
                'success': True, 
                'message': f'تم إنشاء {created_count} ناخب بنجاح',
                'errors': errors
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'خطأ في معالجة الملف: {str(e)}'})
    
    return render(request, 'elections/upload_excel.html', {'type': 'candidate', 'title': 'رفع ملف الناخبين للمرشح'})



# API للحصول على إحصائيات الناخبين
@csrf_exempt
def get_voter_stats(request):
    if request.method == 'GET':
        # منطق الحصول على الإحصائيات
        stats = {
            'total_voters': Voter.objects.count(),
            'voted': Voter.objects.filter(voting_status='voted').count(),
            'not_voted': Voter.objects.filter(voting_status='not_voted').count(),
        }
        return JsonResponse(stats)

# جلب الركائز للمرشح المحدد
@login_required
def get_pillars_for_candidate(request, candidate_id):
    if request.user.user_type != 'entity':
        return JsonResponse({'error': 'غير مصرح'}, status=403)
    
    try:
        entity = get_object_or_404(Entity, user=request.user)
        candidate = get_object_or_404(Candidate, id=candidate_id, entity=entity)
        pillars = candidate.pillars.all()
        
        pillars_data = [{
            'id': pillar.id,
            'name': pillar.user.full_name
        } for pillar in pillars]
        
        return JsonResponse({'pillars': pillars_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

# لوحة تحكم الإدارة
@login_required
def admin_dashboard(request):
    if request.user.user_type != 'admin':
        return redirect('elections:login')
    
    # إحصائيات شاملة
    total_entities = Entity.objects.count()
    total_candidates = Candidate.objects.count()
    total_pillars = Pillar.objects.count()
    total_voters = Voter.objects.count()
    voted_count = Voter.objects.filter(voting_status='voted').count()
    not_voted_count = Voter.objects.filter(voting_status='not_voted').count()
    
    # أحدث الكيانات والمرشحين
    recent_entities = Entity.objects.select_related('user').order_by('-user__created_at')[:5]
    recent_candidates = Candidate.objects.select_related('user', 'entity').order_by('-user__created_at')[:5]
    
    context = {
        'total_entities': total_entities,
        'total_candidates': total_candidates,
        'total_pillars': total_pillars,
        'total_voters': total_voters,
        'voted_count': voted_count,
        'not_voted_count': not_voted_count,
        'voting_percentage': round((voted_count / total_voters * 100) if total_voters > 0 else 0, 2),
        'recent_entities': recent_entities,
        'recent_candidates': recent_candidates,
    }
    return render(request, 'elections/admin_dashboard.html', context)

# إنشاء كيان جديد
@login_required
def create_entity(request):
    if request.user.user_type != 'admin':
        return redirect('elections:login')
    
    if request.method == 'POST':
        form = EntityForm(request.POST, request.FILES)
        if form.is_valid():
            # التحقق من عدم وجود اسم المستخدم مسبقاً
            if CustomUser.objects.filter(username=form.cleaned_data['username']).exists():
                form.add_error('username', 'اسم المستخدم موجود مسبقاً')
            else:
                entity = form.save()
                messages.success(request, f'تم إنشاء الكيان "{entity.entity_name}" بنجاح')
                return redirect('elections:admin_dashboard')
    else:
        form = EntityForm()
    
    return render(request, 'elections/create_entity.html', {'form': form})

# إنشاء مرشح جديد
@login_required
def create_candidate(request):
    if request.user.user_type != 'admin':
        return redirect('elections:login')
    
    entities = Entity.objects.all()
    
    if request.method == 'POST':
        # إنشاء مستخدم جديد
        username = request.POST.get('username')
        password = request.POST.get('password')
        full_name = request.POST.get('full_name')
        phone_number = request.POST.get('phone_number', '')
        entity_id = request.POST.get('entity_id')
        
        if CustomUser.objects.filter(username=username).exists():
            messages.error(request, 'اسم المستخدم موجود مسبقاً')
        else:
            # إنشاء المستخدم
            user = CustomUser.objects.create_user(
                username=username,
                password=password,
                full_name=full_name,
                phone_number=phone_number,
                user_type='candidate'
            )
            
            # إنشاء المرشح
            entity = get_object_or_404(Entity, id=entity_id)
            Candidate.objects.create(
                user=user,
                entity=entity
            )
            
            messages.success(request, f'تم إنشاء المرشح "{full_name}" بنجاح')
            return redirect('elections:admin_dashboard')
    
    context = {'entities': entities}
    return render(request, 'elections/create_candidate.html', context)

# إدارة الكيانات
@login_required
def manage_entities(request):
    if request.user.user_type != 'admin':
        return redirect('elections:login')
    
    entities = Entity.objects.select_related('user').annotate(
        candidates_count=Count('candidates')
    ).order_by('-user__created_at')
    
    # البحث
    search_query = request.GET.get('search', '')
    if search_query:
        entities = entities.filter(
            Q(entity_name__icontains=search_query) |
            Q(user__full_name__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )
    
    # التصفح
    paginator = Paginator(entities, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'elections/manage_entities.html', context)

# إدارة المرشحين
@login_required
def manage_candidates(request):
    if request.user.user_type != 'admin':
        return redirect('elections:login')
    
    candidates = Candidate.objects.select_related('user', 'entity').annotate(
        pillars_count=Count('pillars')
    ).order_by('-user__created_at')
    
    # البحث
    search_query = request.GET.get('search', '')
    if search_query:
        candidates = candidates.filter(
            Q(user__full_name__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(entity__entity_name__icontains=search_query)
        )
    
    # التصفح
    paginator = Paginator(candidates, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'elections/manage_candidates.html', context)

# تحرير كيان
@login_required
def edit_entity(request, entity_id):
    if request.user.user_type != 'admin':
        return redirect('elections:login')
    
    entity = get_object_or_404(Entity, id=entity_id)
    
    if request.method == 'POST':
        # تحديث بيانات المستخدم
        entity.user.full_name = request.POST.get('full_name')
        entity.user.phone_number = request.POST.get('phone_number', '')
        entity.user.save()
        
        # تحديث اسم الكيان
        entity.entity_name = request.POST.get('entity_name')
        entity.save()
        
        messages.success(request, f'تم تحديث الكيان "{entity.entity_name}" بنجاح')
        return redirect('elections:manage_entities')
    
    context = {'entity': entity}
    return render(request, 'elections/edit_entity.html', context)

# حذف كيان
@login_required
def delete_entity(request, entity_id):
    if request.user.user_type != 'admin':
        return redirect('elections:login')
    
    entity = get_object_or_404(Entity, id=entity_id)
    
    if request.method == 'POST':
        entity_name = entity.entity_name
        # حذف المستخدم المرتبط بالكيان (سيحذف الكيان تلقائياً)
        entity.user.delete()
        
        messages.success(request, f'تم حذف الكيان "{entity_name}" بنجاح')
        return redirect('elections:manage_entities')
    
    context = {'entity': entity}
    return render(request, 'elections/delete_entity.html', context)

# تحرير مرشح
@login_required
def edit_candidate(request, candidate_id):
    if request.user.user_type != 'admin':
        return redirect('elections:login')
    
    candidate = get_object_or_404(Candidate, id=candidate_id)
    entities = Entity.objects.all()
    
    if request.method == 'POST':
        # تحديث بيانات المستخدم
        candidate.user.full_name = request.POST.get('full_name')
        candidate.user.phone_number = request.POST.get('phone_number', '')
        candidate.user.save()
        
        # تحديث الكيان المرتبط
        entity_id = request.POST.get('entity_id')
        candidate.entity = get_object_or_404(Entity, id=entity_id)
        candidate.save()
        
        messages.success(request, f'تم تحديث المرشح "{candidate.user.full_name}" بنجاح')
        return redirect('elections:manage_candidates')
    
    context = {
        'candidate': candidate,
        'entities': entities
    }
    return render(request, 'elections/edit_candidate.html', context)

# حذف مرشح
@login_required
def delete_candidate(request, candidate_id):
    if request.user.user_type != 'admin':
        return redirect('elections:login')
    
    candidate = get_object_or_404(Candidate, id=candidate_id)
    
    if request.method == 'POST':
        candidate_name = candidate.user.full_name
        # حذف المستخدم المرتبط بالمرشح (سيحذف المرشح تلقائياً)
        candidate.user.delete()
        
        messages.success(request, f'تم حذف المرشح "{candidate_name}" بنجاح')
        return redirect('elections:manage_candidates')
    
    context = {'candidate': candidate}
    return render(request, 'elections/delete_candidate.html', context)

# صفحة تفاصيل الإحصائيات
@login_required
def statistics_detail(request, stat_type):
    # التحقق من صلاحيات المستخدم
    if request.user.user_type not in ['admin', 'entity', 'candidate', 'pillar']:
        messages.error(request, 'ليس لديك صلاحية للوصول لهذه الصفحة')
        return redirect('elections:login')
    
    context = {
        'stat_type': stat_type,
        'user_type': request.user.user_type,
        'page_title': 'تفاصيل الإحصائيات'
    }
    
    # تحديد البيانات حسب نوع الإحصائية ونوع المستخدم
    if stat_type == 'entities':
        if request.user.user_type == 'admin':
            entities = Entity.objects.all().annotate(
                candidates_count=Count('candidates'),
                voters_count=Count('candidates__voters')
            )
            context.update({
                'entities': entities,
                'page_title': 'تفاصيل الكيانات',
                'total_count': entities.count()
            })
        else:
            messages.error(request, 'ليس لديك صلاحية لعرض هذه البيانات')
            return redirect('elections:login')
    
    elif stat_type == 'candidates':
        if request.user.user_type == 'admin':
            candidates = Candidate.objects.all().select_related('entity').annotate(
                pillars_count=Count('pillars'),
                voters_count=Count('voters')
            )
        elif request.user.user_type == 'entity':
            candidates = Candidate.objects.filter(entity=request.user.entity_profile).annotate(
                pillars_count=Count('pillars'),
                voters_count=Count('voters')
            )
        else:
            messages.error(request, 'ليس لديك صلاحية لعرض هذه البيانات')
            return redirect('elections:login')
        
        context.update({
            'candidates': candidates,
            'page_title': 'تفاصيل المرشحين',
            'total_count': candidates.count()
        })
    
    elif stat_type == 'pillars':
        if request.user.user_type == 'admin':
            pillars = Pillar.objects.all().select_related('candidate', 'candidate__entity').annotate(
                voters_count=Count('voters')
            )
        elif request.user.user_type == 'entity':
            pillars = Pillar.objects.filter(candidate__entity=request.user.entity_profile).select_related('candidate').annotate(
                voters_count=Count('voters')
            )
        elif request.user.user_type == 'candidate':
            pillars = Pillar.objects.filter(candidate=request.user.candidate_profile).annotate(
                voters_count=Count('voters')
            )
        else:
            messages.error(request, 'ليس لديك صلاحية لعرض هذه البيانات')
            return redirect('elections:login')
        
        context.update({
            'pillars': pillars,
            'page_title': 'تفاصيل الركائز',
            'total_count': pillars.count()
        })
    
    elif stat_type == 'voters':
        if request.user.user_type == 'admin':
            voters = Voter.objects.all().select_related('pillar', 'pillar__candidate', 'pillar__candidate__entity')
        elif request.user.user_type == 'entity':
            voters = Voter.objects.filter(pillar__candidate__entity=request.user.entity_profile).select_related('pillar', 'pillar__candidate')
        elif request.user.user_type == 'candidate':
            voters = Voter.objects.filter(pillar__candidate=request.user.candidate_profile).select_related('pillar')
        elif request.user.user_type == 'pillar':
            voters = Voter.objects.filter(pillar=request.user.pillar_profile)
        
        # تطبيق التصفية حسب حالة التصويت إذا تم تمريرها
        vote_status = request.GET.get('vote_status')
        if vote_status == 'voted':
            voters = voters.filter(voting_status='voted')
            context['filter_status'] = 'صوتوا'
        elif vote_status == 'not_voted':
            voters = voters.filter(voting_status='not_voted')
            context['filter_status'] = 'لم يصوتوا'
        
        # تطبيق البحث والفلترة
        search_query = request.GET.get('search')
        district_filter = request.GET.get('district')
        sub_district_filter = request.GET.get('sub_district')
        
        if search_query:
            voters = voters.filter(
                    Q(name__icontains=search_query) |
                    Q(voter_number__icontains=search_query) |
                    Q(phone_number__icontains=search_query)
                )
            context['search_query'] = search_query
        
        if district_filter:
            voters = voters.filter(district=district_filter)
            context['district_filter'] = district_filter
        
        if sub_district_filter:
            voters = voters.filter(sub_district=sub_district_filter)
            context['sub_district_filter'] = sub_district_filter
        
        # التصفح
        paginator = Paginator(voters, 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # الحصول على قوائم المناطق والنواحي المتاحة
        all_voters_for_filters = voters.model.objects.all()
        if request.user.user_type == 'entity':
            all_voters_for_filters = voters.model.objects.filter(pillar__candidate__entity=request.user.entity_profile)
        elif request.user.user_type == 'candidate':
            all_voters_for_filters = voters.model.objects.filter(pillar__candidate=request.user.candidate_profile)
        elif request.user.user_type == 'pillar':
            all_voters_for_filters = voters.model.objects.filter(pillar=request.user.pillar_profile)
        
        districts = all_voters_for_filters.values_list('district', flat=True).distinct().order_by('district')
        sub_districts = all_voters_for_filters.values_list('sub_district', flat=True).distinct().order_by('sub_district')
        
        context.update({
            'voters': page_obj,
            'page_title': 'تفاصيل الناخبين',
            'total_count': voters.count(),
            'voted_count': voters.filter(voting_status='voted').count(),
            'not_voted_count': voters.filter(voting_status='not_voted').count(),
            'districts': districts,
            'sub_districts': sub_districts,
        })
    
    elif stat_type == 'voted':
        # إحصائيات المصوتين
        if request.user.user_type == 'admin':
            voters = Voter.objects.filter(voting_status='voted').select_related('pillar', 'pillar__candidate', 'pillar__candidate__entity')
        elif request.user.user_type == 'entity':
            voters = Voter.objects.filter(voting_status='voted', pillar__candidate__entity=request.user.entity_profile).select_related('pillar', 'pillar__candidate')
        elif request.user.user_type == 'candidate':
            voters = Voter.objects.filter(voting_status='voted', pillar__candidate=request.user.candidate_profile).select_related('pillar')
        elif request.user.user_type == 'pillar':
            voters = Voter.objects.filter(voting_status='voted', pillar=request.user.pillar_profile)
        
        # التصفح
        paginator = Paginator(voters, 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context.update({
            'voters': page_obj,
            'page_title': 'الناخبون الذين صوتوا',
            'total_count': voters.count()
        })
    
    elif stat_type == 'not_voted':
        # إحصائيات غير المصوتين
        if request.user.user_type == 'admin':
            voters = Voter.objects.filter(voting_status='not_voted').select_related('pillar', 'pillar__candidate', 'pillar__candidate__entity')
        elif request.user.user_type == 'entity':
            voters = Voter.objects.filter(voting_status='not_voted', pillar__candidate__entity=request.user.entity_profile).select_related('pillar', 'pillar__candidate')
        elif request.user.user_type == 'candidate':
            voters = Voter.objects.filter(voting_status='not_voted', pillar__candidate=request.user.candidate_profile).select_related('pillar')
        elif request.user.user_type == 'pillar':
            voters = Voter.objects.filter(voting_status='not_voted', pillar=request.user.pillar_profile)
        
        # التصفح
        paginator = Paginator(voters, 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context.update({
            'voters': page_obj,
            'page_title': 'الناخبون الذين لم يصوتوا',
            'total_count': voters.count()
        })
    
    elif stat_type == 'updated':
        # إحصائيات المحدثين
        if request.user.user_type == 'admin':
            voters = Voter.objects.filter(card_status='updated').select_related('pillar', 'pillar__candidate', 'pillar__candidate__entity')
        elif request.user.user_type == 'entity':
            voters = Voter.objects.filter(card_status='updated', pillar__candidate__entity=request.user.entity_profile).select_related('pillar', 'pillar__candidate')
        elif request.user.user_type == 'candidate':
            voters = Voter.objects.filter(card_status='updated', pillar__candidate=request.user.candidate_profile).select_related('pillar')
        elif request.user.user_type == 'pillar':
            voters = Voter.objects.filter(card_status='updated', pillar=request.user.pillar_profile)
        
        # التصفح
        paginator = Paginator(voters, 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context.update({
            'voters': page_obj,
            'page_title': 'الناخبون المحدثون',
            'total_count': voters.count()
        })
    
    elif stat_type == 'not_updated':
        # إحصائيات غير المحدثين
        if request.user.user_type == 'admin':
            voters = Voter.objects.filter(card_status='not_updated').select_related('pillar', 'pillar__candidate', 'pillar__candidate__entity')
        elif request.user.user_type == 'entity':
            voters = Voter.objects.filter(card_status='not_updated', pillar__candidate__entity=request.user.entity_profile).select_related('pillar', 'pillar__candidate')
        elif request.user.user_type == 'candidate':
            voters = Voter.objects.filter(card_status='not_updated', pillar__candidate=request.user.candidate_profile).select_related('pillar')
        elif request.user.user_type == 'pillar':
            voters = Voter.objects.filter(card_status='not_updated', pillar=request.user.pillar_profile)
        
        # التصفح
        paginator = Paginator(voters, 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context.update({
            'voters': page_obj,
            'page_title': 'الناخبون غير المحدثين',
            'total_count': voters.count()
        })
    
    else:
        messages.error(request, 'نوع الإحصائية غير صحيح')
        return redirect('elections:admin_dashboard')
    
    return render(request, 'elections/statistics_detail.html', context)

@login_required
def appearance_settings(request):
    """إعدادات المظهر - متاحة للمسؤولين فقط"""
    if request.user.user_type != 'admin':
        messages.error(request, 'غير مسموح لك بالوصول إلى هذه الصفحة')
        return redirect('elections:admin_dashboard')
    
    current_settings = AppearanceSettings.get_active_settings()
    
    if request.method == 'POST':
        primary_color = request.POST.get('primary_color', '#007bff')
        secondary_color = request.POST.get('secondary_color', '#6c757d')
        button_text_color = request.POST.get('button_text_color', '#ffffff')
        card_title_color = request.POST.get('card_title_color', '#212529')
        
        # التحقق من صحة الألوان (hex format)
        if not (primary_color.startswith('#') and len(primary_color) == 7):
            messages.error(request, 'اللون الأساسي الأول غير صحيح')
            return render(request, 'elections/appearance_settings.html', {
                'current_settings': current_settings
            })
        
        if not (secondary_color.startswith('#') and len(secondary_color) == 7):
            messages.error(request, 'اللون الأساسي الثاني غير صحيح')
            return render(request, 'elections/appearance_settings.html', {
                'current_settings': current_settings
            })
            
        if not (button_text_color.startswith('#') and len(button_text_color) == 7):
            messages.error(request, 'لون نص الأزرار غير صحيح')
            return render(request, 'elections/appearance_settings.html', {
                'current_settings': current_settings
            })
        
        # إنشاء إعدادات جديدة
        new_settings = AppearanceSettings.objects.create(
            primary_color=primary_color,
            secondary_color=secondary_color,
            button_text_color=button_text_color,
            card_title_color=card_title_color,
            is_active=True
        )
        
        messages.success(request, 'تم حفظ إعدادات المظهر بنجاح')
        return redirect('elections:appearance_settings')
    
    return render(request, 'elections/appearance_settings.html', {
        'current_settings': current_settings
    })
